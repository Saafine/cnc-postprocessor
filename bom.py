from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment
import translate


def excel_boards(furniture_thickness, texture, backwall_texture, wall, board, columns, shelf_array, door, sockle, divider, tabletop, order_id, order_folder):
    # standard values
    wb = Workbook()
    dest_filename = order_folder+'plyty.xlsx'
    ws1 = wb.active  # referencing : ws1['A1'] = 3.14 or ws1.cell(row=1, column=1, value='3.14')
    ws1.title = "Plyty"
    ws1.sheet_properties.pageSetUpPr.fitToPage = True  # print -> fit all columns on 1 page
    ws1.page_setup.fitToHeight = False  # print -> fit all columns on 1 page

    # create column titles
    ws1.cell(row=1, column=1, value=order_id)  # set order id
    column_names = ['Element', 'Ilość sztuk', 'Długość brutto', 'Szerokość brutto', 'Materiał', 'Wykończenie', 'Symbol', 'Długość netto', 'Szerokość netto']
    col_names = len(column_names)+1

    row = 2
    for i, name in enumerate(column_names):
        ws1.cell(row=row, column=i+1, value=name)
    row += 1
    get_mat_info = translate.material(furniture_thickness, texture)  # translate texture name
    texture = get_mat_info[0]
    material = get_mat_info[1]
    material_excess = get_mat_info[2]
    sockle_thickness = 18
    # material2 = 'MDF18'
    # material3 = 'MDF12'

    # append wall info to excel
    for run_col in range(len(wall)):
        wall_info = ['ściana', 1, wall[run_col].height+material_excess, wall[run_col].depth+material_excess, material, texture, 'WA'+str(run_col+1), wall[run_col].height, wall[run_col].depth]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=wall_info[column-1])
        row += 1

    # append board info to excel
    for run_col in range(columns):
        col_BN = [0, 0, 0]
        col_BD = [0, 0, 0]
        for shelf in range(1, shelf_array[run_col]+1):
            if board[run_col][shelf].recess:
                col_BN[0] += 1
                col_BN[1] = board[run_col][shelf].width
                col_BN[2] = board[run_col][shelf].depth
            else:
                col_BD[0] += 1
                col_BD[1] = board[run_col][shelf].width
                col_BD[2] = board[run_col][shelf].depth
        if col_BN[0] > 0:
            board_info = ['półka', col_BN[0], col_BN[1]+material_excess, col_BN[2]+material_excess, material, texture, 'BN' + str(run_col + 1), col_BN[1], col_BN[2]]
            for column in range(1, col_names):
                if col_BN[0] > 1:
                        ws1.cell(row=row, column=column).fill = PatternFill("solid", fgColor="ffdfdf")
                ws1.cell(row=row, column=column, value=board_info[column-1])
            row += 1
        if col_BD[0] > 0:
            board_info = ['półka', col_BD[0], col_BD[1]+material_excess, col_BD[2]+material_excess, material, texture, 'BD' + str(run_col + 1), col_BD[1], col_BD[2]]
            for column in range(1, col_names):
                if col_BD[0] > 1:
                        ws1.cell(row=row, column=column).fill = PatternFill("solid", fgColor="ffdfdf")
                ws1.cell(row=row, column=column, value=board_info[column - 1])
            row += 1

    # append door info to excel
    for run_door in range(len(door)):
        if door[run_door].doortype == 'door':
            doortype = 'drzwi'
        elif door[run_door].doortype == 'drawer':
            doortype = 'czoło szuflady'
        elif door[run_door].doortype == 'flap':
            doortype = 'klapa'
        elif door[run_door].doortype == 'doubleLEFT':
            doortype = 'drzwi lewe'
        elif door[run_door].doortype == 'doubleRIGHT':
            doortype = 'drzwi prawe'
        else:
            doortype = ''

        get_door_info = translate.material(furniture_thickness, door[run_door].texture)
        get_door_excess = get_door_info[2]

        door_info = [doortype, 1, door[run_door].height+get_door_excess, door[run_door].width+get_door_excess, get_door_info[1], get_door_info[0], 'D' + str(run_door + 1), door[run_door].height, door[run_door].width]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=door_info[column - 1])
        row += 1
        # if door[run_door].doortype == 'drawer':
        #     side_info = ['bok/bok szuflady', 2, door[run_door].drawer_side_depth+material_excess, door[run_door].drawer_side_height+material_excess, material3, texture, 'DS' + str(run_door + 1)]
        #     for column in range(1, col_names):
        #         ws1.cell(row=row, column=column, value=side_info[column - 1])
        #         ws1.cell(row=row, column=column).fill = PatternFill("solid", fgColor="ffdfdf")
        #     row += 1
        #     side_info = ['przod/tyl szuflady', 2, door[run_door].drawer_front_width+material_excess, door[run_door].drawer_front_height+material_excess, material3, texture, 'DF' + str(run_door + 1)]
        #     for column in range(1, col_names):
        #         ws1.cell(row=row, column=column, value=side_info[column - 1])
        #         ws1.cell(row=row, column=column).fill = PatternFill("solid", fgColor="ffdfdf")
        #     row += 1

    # append sockle info to excel
    for run_sockle in range(len(sockle)):
        sockle_info = ['cokół', 1, sockle[run_sockle].width, sockle[run_sockle].height, sockle_thickness, texture, 'S' + str(run_sockle + 1), sockle[run_sockle].width, sockle[run_sockle].height]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=sockle_info[column - 1])
        ws1.cell(row=row, column=10, value='NETTO')
        row += 1

    for run_divider in range(len(divider)):
        divider_info = ['pionik', 1, divider[run_divider].height+material_excess, divider[run_divider].depth+material_excess, material, texture, 'E' + str(run_divider + 1), divider[run_divider].height, divider[run_divider].depth]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=divider_info[column - 1])
        row += 1

    for run_tabletop in range(len(tabletop)):
        tabletop_info = ['blat', 1, tabletop[run_tabletop].width+material_excess, tabletop[run_tabletop].depth+material_excess, material, texture, 'BT' + str(run_tabletop + 1), tabletop[run_tabletop].width, tabletop[run_tabletop].depth]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=tabletop_info[column - 1])
        row += 1

    for run_btm in range(columns):
        btm_info = ['dno', 1, board[run_btm][0].width+material_excess, board[run_btm][0].depth+material_excess, material, texture, 'BS' + str(run_btm + 1), board[run_btm][0].width, board[run_btm][0].depth]
        for column in range(1, col_names):
            ws1.cell(row=row, column=column, value=btm_info[column - 1])
        row += 1

    # Styling begins
    border_color = 'FF000000'
    border_style = 'thin'
    border_top = Border(top=Side(border_style=border_style, color=border_color),
                                  right=Side(border_style=border_style, color=border_color),
                                  bottom=Side(border_style=border_style, color=border_color),
                                  left=Side(border_style=border_style, color=border_color))
    for c in range(1, col_names):
        for r in range(1, row):
            ws1.cell(row=r, column=c).border = border_top

    fill_me = PatternFill("solid", fgColor="d8e8ff")
    center_me = Alignment(horizontal="center")
    for c in range(1, col_names):
        ws1.cell(row=2, column=c).fill = fill_me
        ws1.cell(row=2, column=c).alignment = center_me

    ws1.column_dimensions["A"].width = 17
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 9
    ws1.column_dimensions["H"].width = 15
    ws1.column_dimensions["I"].width = 15

    wb.save(dest_filename)  # save file
