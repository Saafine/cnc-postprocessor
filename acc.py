#!/usr/bin/env python
# -*- coding: utf-8 -*-
# enable UTF-8 characters in the console

from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment


class Used:
    pin = {  # for: tabletops and dividers
        'name': 'trzpień mimośrodowy',
        'producer': 'titus',
        'price': '2',
        'code': 'TITUS123',
        'amount': 0
    }
    cam = {
        'name': 'mimośród',
        'producer': 'titus',
        'price': '2',
        'code': 'TITUS123',
        'amount': 0
    }
    dowel = {
        'name': 'kołek',
        'producer': 'meblopol',
        'price': '2',
        'code': '',
        'amount': 0
    }
    brd = {  # for: tabletops and boards
        'name': 'RAFIX, złaczka do półek',
        'producer': 'Hafele',
        'price': '2',
        'code': '263.11.705',
        'amount': 0
    }

    wall_brd = {  # for: tabletops and boards
        'name': 'RAFIX, trzpień na półki',
        'producer': 'Hafele',
        'price': '2',
        'code': '263.20.131',
        'amount': 0
    }

    btm = {
        'name': 'Tab-18, złączka do den',
        'producer': 'Hafele',
        'price': '2',
        'code': '263.64.536',
        'amount': 0
    }

    wall_btm = {
        'name': 'Tab-18, mocowanie na dno',
        'producer': 'Hafele',
        'price': '2',
        'code': '263.60.538',
        'amount': 0
    }

    door_hinge_straight = {
        'name': 'zawias prosty',
        'producer': 'GTV',
        'price': '2',
        'code': '',
        'amount': 0
    }

    door_hinge_110 = {
        'name': 'zawias 110 stopni',
        'producer': 'GTV',
        'price': '2',
        'code': '',
        'amount': 0
    }

    door_plate = {
        'name': 'prowadnik do zawiasu',
        'producer': 'GTV',
        'price': '2',
        'code': '',
        'amount': 0
    }

    door_hinge_flap = {
        'name': 'zawias do klap',
        'producer': 'Hafele',
        'price': '2',
        'code': '',
        'amount': 0
    }

    door_arm_flap = {
        'name': 'ramię do klap',
        'producer': 'Hafele',
        'price': '2',
        'code': '',
        'amount': 0
    }

    door_push = {
        'name': 'wypychacz',
        'producer': 'TITUS',
        'price': '2',
        'code': '',
        'amount': 0
    }

    drawer_slider = {
        'name': 'prowadnica, dł. ',
        'length': 0,
        'producer': 'BLUM',
        'price': '2',
        'code': '',
        'amount': 0
    }

    drawer_slider_sync = {
        'name': 'synchronizator do szuflady',
        'producer': 'BLUM',
        'price': '2',
        'code': '',
        'amount': 0
    }

    drawer_slider_push = {
        'name': 'tip-on do szuflady',
        'producer': 'BLUM',
        'price': '2',
        'code': '',
        'amount': 0
    }

    hanger_tube = {
        'name': 'wieszak - drążek',
        'producer': '?',
        'price': '2',
        'code': '',
        'amount': 0
    }

    hanger_runner = {
        'name': 'wieszak - wysuwany',
        'producer': 'HAFELE',
        'price': '2',
        'code': '',
        'amount': 0
    }

    btm_sockle = {
        'name': 'Keku, część mocowana do dna',
        'producer': 'Hafele',
        'price': '2',
        'code': '',
        'amount': 0
    }

    sockle = {
        'name': 'Keku, część mocowana do cokołu',
        'producer': 'Hafele',
        'price': '2',
        'code': '',
        'amount': 0
    }


def excel_accs(order_id, order_folder):
    # standard values
    wb = Workbook()
    dest_filename = order_folder+'akcesoria.xlsx'
    ws1 = wb.active  # referencing : ws1['A1'] = 3.14 or ws1.cell(row=1, column=1, value='3.14')
    ws1.title = "Akcesoria"
    ws1.sheet_properties.pageSetUpPr.fitToPage = True  # print -> fit all columns on 1 page
    ws1.page_setup.fitToHeight = False  # print -> fit all columns on 1 page

    # create column titles
    ws1.cell(row=1, column=1, value=order_id)  # set order id
    column_names = ['Element', 'Ilość sztuk', 'Producent', 'Kod', 'Cena netto']
    col_names = len(column_names) + 1

    row = 2
    for i, name in enumerate(column_names):
        ws1.cell(row=row, column=i+1, value=name)
    row += 1

    def export_data(object, row):
        if object['amount'] == 0:
            return 0
        for column in range(1, col_names):
            cell_info = [object['name'], object['amount'], object['producer'], object['code'], object['price']]
            ws1.cell(row=row, column=column, value=cell_info[column-1])

    accs = [Used.pin, Used.cam, Used.dowel, Used.brd, Used.wall_brd, Used.btm, Used.wall_btm,
            Used.door_hinge_straight, Used.door_hinge_110, Used.door_plate, Used.door_hinge_flap, Used.door_arm_flap, Used.door_push, Used.drawer_slider_sync,
            Used.drawer_slider_push, Used.hanger_tube, Used.hanger_runner, Used.btm_sockle, Used.sockle]
    for each in accs:
        data = export_data(each, row)
        if data != 0:  # if amount of specific accessory equals 0, skip appending a row
            row += 1

    for column in range(1, col_names):
        cell_info = [Used.drawer_slider['name'] + str(Used.drawer_slider['length']), Used.drawer_slider['amount'], Used.drawer_slider['producer'], Used.drawer_slider['code'], Used.drawer_slider['price']]
        if Used.drawer_slider['amount'] != 0:
            ws1.cell(row=row, column=column, value=cell_info[column-1])
    row += 1

    # Styling begins
    border_color = 'FF000000'
    border_style = 'thin'
    border_top = Border(top=Side(border_style=border_style, color=border_color),
                                  right=Side(border_style=border_style, color=border_color),
                                  bottom=Side(border_style=border_style, color=border_color),
                                  left=Side(border_style=border_style, color=border_color))
    for c in range(1, col_names):
        for r in range(1, row):
            ws1.cell(row=r, column=c).border = border_top

    fill_me = PatternFill("solid", fgColor="d8e8ff")
    center_me = Alignment(horizontal="center")
    for c in range(1, col_names):
        ws1.cell(row=2, column=c).fill = fill_me
        ws1.cell(row=2, column=c).alignment = center_me

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14
    wb.save(dest_filename)  # save file
